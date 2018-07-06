# -*- coding:utf-8 -*-
import sys
import os
import signal
import pexpect
import re
import time
from ipaddress import IPv4Network
from pymongo import MongoClient
from datetime import datetime
import paramiko
from fpdf import FPDF, HTMLMixin
import xlrd
from openpyxl.workbook import Workbook
import requests
from requests.auth import HTTPBasicAuth
from requests_ntlm import HttpNtlmAuth
from ciscoconfparse import CiscoConfParse
import xlsxwriter
from bs4.element import Tag
from bs4.element import NavigableString
from bs4 import BeautifulSoup
from django.conf import settings
from io import StringIO, BytesIO

import shutil

CONNECTION_PROBLEM = -1
INVALID_HOSTNAME = 0
INVALID_AUTH = 1
INVALID_COMMAND = 2

local_id_regex = {
            'MEXICO': "(A|C|D|F|S)([B|T]\d{1}|\d{2})-\d{4}-\d{4}",
            'BRASIL': "(HORT|CSL|CEM|SBS|MBB|LDA|MGA|SJC|SOO|GNA|SGO|PAS|PGE|RPO|ULA|GRS|SPO|PAE|CAS|RJO|SJP|SDR|BPI|ITU|CTA|MNS|SOC|CBO|MCO|UPS|FLA|JBO|SRR|PSO|RCO|LNS|VGA|ABS|CSC|YCA|ETA|JVE|LNA|BRU|JAI|SPA|AMA|STB|NTL|MCL|CBM|NHO|PLT|BRE|LGS|GAMA|PTA|FSA|PTS|SBO|CE|FLA|PWM)\/(IP|MULTI|FAST)\/(\d{5})",
            'PERU': "\d{7}",
            'HONDURAS': "\d{6}",
            'GUATEMALA': "\d{8}",
            'CHILE': "\d{2}-\d{2}-\d{10}",
            'COLOMBIA': "\w{3}\d{4}",
            'EL SALVADOR': "IP\d{7}",
            'ESTADOS UNIDOS': 'ITFS\-\d{11}',
            'ARGENTINA': '\d{7}',
            'VENEZUELA': '',
            'ECUADOR': "asd",
            'AUSTRIA': "",
            'PARAGUAY': "",
            'NICARAGUA': "",
            'ESPAÑA': "",
            "URUGUAY": "",
            'BOLIVIA': "",
            'BERMUDAS': "",
            'CANADA': "",
            'TRINIDAD Y TOBAGO': "",
            'REPUBLICA DOMINICANA': "^\d{3}-\d{3}-\d{4}$",
            'COSTA RICA': "",
        }

imp_list = [
    'GASTON GUTIERREZ',
    'DIEGO MARTIN CAPELLO',
    'GUILLERMO DIAZ'
]

providers = {
    'MEXICO': 'TELMEX',
    'BRASIL': 'EMBRATEL',
    'ESTADOS UNIDOS': 'TELMEX USA',
}

telmex_as_list = {
    'ESTADOS UNIDOS': '28513',
    'MEXICO': '8151',
    'BRASIL': '4230',
    'COSTA RICA': '14754',
    'GUATEMALA': '14754',
    'NICARAGUA': '14754',
    'HONDURAS': '14754',
    'COLOMBIA': '14080',
    'PERU': '12252',
    'ECUADOR': '23487',
    'CHILE': '100',
    'ARGENTINA': '11664',
    }

colo_host_to_ip = {
    'A9KBUCARAMANGA': '10.10.66.236',
    'GSANDIEGO': '10.10.66.199',
}

hostname_test_brasil = 'GACC01.FNS' # IOS
hostname_test_mexico = 'vpn-yuc-plaza-32' # XR

def getProvider(link):
    return providers[link.country] if providers.get(link.country, 0) else 'CLARO'


def countCountry(list, country):
    if country not in [safe_list_get(pending, 0, 0) for pending in list]:
        list.append([country, 1])
    else:
        for element in list:
            if safe_list_get(element, 0, 0) == country:
                element[1] += 1


def printProgress (iteration, total, prefix = '', suffix = '', decimals = 1, barLength = 100):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        barLength   - Optional  : character length of bar (Int)
    """
    formatStr       = "{0:." + str(decimals) + "f}"
    percents        = formatStr.format(100 * (iteration / float(total)))
    filledLength    = int(round(barLength * iteration / float(total)))
    bar             = '█' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percents, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()

def is_xr(config):
    p = re.compile("(IOS XR)")
    if p.search(config):
        return True

def routerIOSFromConfig(config):
    if "IOS XR" in config:
        return "xr"
    elif "juniper" in config:
        return "junos"
    else:
        return "ios"

def safe_list_get(l, idx, default):
      try:
        return l[idx]
      except IndexError:
        return default

def downloadPepsicoTracker(path):
    url = "http://telmexusaintra/sites/CustomerCare/_layouts/download.aspx?SourceUrl=%2Fsites%2FCustomerCare%2FSpecial%20Reports%2FPepsico%20Source%2FPEPSICO%20Source%20Internal%20v3%2Exlsx&Source=http%3A%2F%2Ftelmexusaintra%2Fsites%2FCustomerCare%2FSpecial%2520Reports%2FForms%2FAllItems%2Easpx%3FRootFolder%3D%252fsites%252fCustomerCare%252fSpecial%2520Reports%252fPepsico%2520Source%26FolderCTID%3D%26View%3D%257b3C3015E1%252d486B%252d499C%252dB1A5%252dED2ED111A595%257d&FldUrl="
    username = "telmex-usa\\guillermo.diaz"
    password = "Motorola1987$"
    r = requests.get(url, auth=HttpNtlmAuth(username, password), stream=True)
    print('status' + str(r.status_code))
    if r.status_code == 200:
        with open(path, 'wb') as f:
            r.raw.decode_content = True
            shutil.copyfileobj(r.raw, f)

    #return r.

def get_info_for_site_name(nsr, column, link_type='Primary'):
    info = str()
    path = "%s/PEPSICO Source Internal v3.xlsx" % settings.BASE_DIR
    columns = {
        "site_name": 0,
        "as": 3,
        "local_id": 9,
        "nsr": 11,
        "pgla": 12,
        "interface": 19,
        "metal": 25,
        "tier": 26
    }

    #downloadPepsicoTracker(path)

    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    for rownum in range(sh.nrows):
        row_data = sh.row_values(rownum)
        if nsr == str(row_data[columns['nsr']]).strip():
            info = row_data[columns[column]]
            break
    #print(info)
    return info if isinstance(info, str) else str(int(info))

def f(x):
    return {
        'a': 1,
        'b': 2,
    }[x]

brasil_code_for_command = {'show': 'ISynu%7D', 'ping': 'ISvotm', 'traceroute': 'ISzxgi'}

def getToPrompt(chan, path, protocol, username, password, extra=''):
    chan.send(protocol + ' ' + path + ' ' + extra +'\n')
    buff = ''
    while buff.find('login as: ') < 0 and buff.find('Username: ') < 0 and buff.find('password: ') < 0 and buff.find(
            '(yes/no)? ') < 0 and buff.rfind('Connection refused') < 0:
        print('looking for prompt')
        resp = chan.recv(9999)
        buff += resp
    if buff.find('Username: ') != -1:
        print('is user')
        chan.send(username + '\n')
        while not buff.endswith('Password: '):
            print('is password after user')
            resp = chan.recv(9999)
            buff += resp
        chan.send(password + '\n')
    elif buff.find('password: ') != -1:
        print('is password')
        chan.send(password + '\n')
    elif buff.find('(yes/no)? ') != -1:
        print('is encryp keys')
        chan.send("yes\n")
        while not buff.endswith('password: '):
            print('is pass after encryp keys')
            resp = chan.recv(9999)
            buff += resp
        chan.send(password + '\n')
    elif buff.find('Connection refused') == -1:
        return 0

    buff = ''
    while not buff.endswith('#') and buff.find('RP/0/RSP0/CPU0') == -1:
        resp = chan.recv(9999)
        buff += resp

    return 1

def get_output(channel):
    chunk = ''
    time.sleep(2)
    while channel.recv_ready():
        chunk += channel.recv(9999).decode('ISO-8859-1')
        time.sleep(1)
    return chunk

def auth_to(hostname, username, password, chan):
    #print('connecting')
    chan.send('ssh ' + hostname + '\n')
    output = get_output(chan)
    #print(output, 'Connection refused' in output, hostname == '187.128.3.74')
    if 'Connection refused' in output and hostname == '187.128.3.74':
        chan.send('telnet ' + hostname + '\n')
        output = get_output(chan)
    elif 'Invalid input' in output:
        print('CHILE_NET')
        chan.send('telnet ' + hostname + ' /vrf CHILE_NET /source-interface lo0\n')
        output = get_output(chan)
    if 'Connection refused' in output:
        return False
    print('connected')
    if output.find('Connection refused') != -1 or output.find('known_hosts') != -1 or output.find('Connection closed') != -1:
        print('refused')
        return False
    elif output.find('Username: ') != -1:
        print('user')
        chan.send(username + '\n')
        output = get_output(chan)
        print(output)
        if output.find('Password: ') != -1:
            chan.send(password + '\n')
            output = get_output(chan)
            print(output)
            if output.find('Permission denied') != -1:
                return False
        else:
            return False
    elif output.find('password: ') != -1:
        print('pass')
        chan.send(password + '\n')
        output = get_output(chan)
        if output.find('Permission denied') != -1:
            return False
    elif output.find('(yes/no)? ') != -1:
        print('key')
        chan.send("yes\n")
        output = get_output(chan)
        if output.find('Username: ') != -1:
            chan.send(username + '\n')
            output = get_output(chan)
            if output.find('Password: ') != -1:
                chan.send(password + '\n')
                output = get_output(chan)
                if output.find('Permission denied') == -1:
                    return False
            else:
                return False
        elif output.find('password: ') != -1:
            chan.send(password + '\n')
            output = get_output(chan)
            if output.find('Permission denied') == -1:
                return False

    return True

def open_ssh_session(hostname, username, password, port, channel=None):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    ssh.connect(hostname, username=username, password=password, port=port, sock=channel)

    return ssh

def get_config_from(country, hostname, command='show running-config', l=True):
    lg = country.lg
    config = ''
    if not lg:
        return 2
    if lg.protocol == 'url':
        if country.name == 'MEXICO':
            print('Mexico')
            values = {'ip': hostname, 'cmd': command, 'submit': 'Submit'}
            r = requests.post(lg.path, auth=(lg.username, lg.password),
                              data=values, timeout=2000)
            config = r.text

            if 'Invalid&nbsp;command' in config:
                return INVALID_COMMAND
            elif 'Invalid&nbsp;source&nbsp;address' in config:
                return INVALID_HOSTNAME
            elif 'do not have access authorization for account' in config:
                return INVALID_AUTH

            config = re.sub("<font color='\#111111\'>", ' ', config, flags=re.IGNORECASE)
            config = re.sub("</font>", ' ', config, flags=re.IGNORECASE)
            config = re.sub("<td width='100'>", ' ', config, flags=re.IGNORECASE)
            config = re.sub("</td>", ' ', config, flags=re.IGNORECASE)
            config = re.sub("     ", "", config, flags=re.IGNORECASE)
            config = re.sub("&nbsp;", " ", config, flags=re.IGNORECASE)
            config = re.sub("<br>", "\n", config, flags=re.IGNORECASE)

        elif country.name == 'BRASIL':
            print('brasil')
            try:
                response = requests.post(lg.path + '&eqpto=' + hostname, auth=HTTPBasicAuth(lg.username, lg.password), timeout=2000)
            except requests.exceptions.ConnectionError as e:
                return CONNECTION_PROBLEM

            html = response.text

            p = re.compile('&arg[0-2]\=[a-zA-Z0-9_%]*')

            args = []
            for match in p.finditer(html):
                arg = match.group()
                arg = arg[arg.find('=') + 1:]
                args.append(arg)

            command_code = brasil_code_for_command.get(command.split(' ')[0], 0)

            if not command_code:
                return INVALID_COMMAND

            command = command_code + command

            if len(args) == 3:
                url = 'http://10.243.187.170/grb/topologia_rede/www/executar_comandos_telnet_proxy.php?&arg0=' + args[
                    0] + '&arg1=' + args[1] + '&arg2=' + args[2] + '&id_rede=15&comando=' + command
                r = requests.post(url, auth=HTTPBasicAuth(lg.username, lg.password), timeout=2000)
                config = r.text
                config = config[config.find(r"\n") + 2:config.rfind(r"</pre>")]
                config = "\n".join(config.split(r"\n"))
            else:
                return 0

    elif lg.protocol == 'ssh':
        if country.name in ['GUATEMALA', 'HONDURAS', 'NICARAGUA', 'COSTA RICA']:
            child = pexpect.spawn('./vpn_script.sh')
            child.expect('gdiaz: ')
            child.sendline('Wktkm1987#')

        try:
            ssh = open_ssh_session(lg.path, username=lg.username, password=lg.password, port=lg.port)
        except paramiko.ssh_exception.AuthenticationException:
            return INVALID_AUTH
        except TimeoutError:
            return CONNECTION_PROBLEM

        #transport = ssh.get_transport()
        #local_addr = ('127.0.0.1', 10022)
        #dest_addr = (hostname, 22)

        #try:
        #    channel = transport.open_channel('direct-tcpip', dest_addr, local_addr)
        #except paramiko.ssh_exception.ChannelException:
        #    return CONNECTION_PROBLEM

        chan = ssh.invoke_shell()
        get_output(chan)
        creds = lg.credentials.all()

        if lg.lg:
            if auth_to(lg.lg.path, lg.lg.username, lg.lg.password, chan):
                creds = lg.lg.credentials.all()
            else:
                return INVALID_AUTH

        for cred in creds:
            if auth_to(hostname, cred.username, cred.password, chan):
                print('authenticated')
                chan.send('terminal length 0\n')
                get_output(chan)
                print(command)
                chan.send(command + '\n')
                config = get_output(chan)
                print('config downloaded')
                ssh.close()
                break

        else:
            return INVALID_AUTH

    if not isinstance(config, list) and l:
        config = config.split("\n")
    elif isinstance(config, list) and not l:
        config = ''.join(config)

    if country.name in ['GUATEMALA', 'HONDURAS', 'NICARAGUA', 'COSTA RICA']:
        os.killpg(os.getpgid(child.pid), signal.SIGTERM)

    return config


def get_config(country, hostname, command='show running-config', list=True):

    mexico = re.compile("(148|187|189|200|201)\.\d{1,3}\.\d{1,3}\.\d{1,3}")
    brasil = re.compile("[A-Z]+")
    cam    = re.compile("\d{1,3}\.192\.\d{1,3}\.\d{1,3}\s{2}255.255.255.255")

    if country.name == "MEXICO":
        url = 'http://200.33.150.46:2002/lgvpnuninet.php'
        username = 'guillermo.los'
        password = 'CtRxPcmxU0'
        values = {'ip': hostname, 'cmd': command, 'submit': 'Submit'}
        r = requests.post(url, auth=HTTPBasicAuth(username, password),
                                 data=values, timeout=2000)
        config = r.text
        p = re.compile("(Invalid)")
        if p.search(config):
            return 0

        config = re.sub("<font color='\#111111\'>", ' ', config, flags=re.IGNORECASE)
        config = re.sub("</font>", ' ', config, flags=re.IGNORECASE)
        config = re.sub("<td width='100'>", ' ', config, flags=re.IGNORECASE)
        config = re.sub("</td>", ' ', config, flags=re.IGNORECASE)
        config = re.sub("     ", "", config, flags=re.IGNORECASE)
        config = re.sub("&nbsp;", " ", config, flags=re.IGNORECASE)
        config = re.sub("<br>", "\n", config, flags=re.IGNORECASE)
    elif brasil.search(hostname):
        #print('Brazil')
        url = 'http://10.243.187.170/grb/topologia_rede/www/executar_comandos_telnet.php?arg0=ectsq%21u%7C%7C+P%7Cq%25q%7D&id_rede=15&eqpto=' + hostname +'&tcos=1458420918474'
        username = 'dcapello@latam'
        password = 'latam'
        first_command = command.split(' ')[0]

        if first_command == "show":
            command_code = "ISynu%7D"
        elif first_command == "ping":
            command_code = "ISvotm"
        else:
            command_code = "ISzxgi"
        command = command.replace(r':', '%3A')
        command = command.replace(r'|', '%7C')
        command = command.replace(' ', '+')

        try:
            response = requests.post(url, auth=HTTPBasicAuth(username, password), timeout=2000)
        except requests.exceptions.ConnectionError as e:
            return -1 # No connection to Looking Glass

        html = response.text
        #print(html)
        args = []

        p = re.compile('&arg[0-2]\=[a-zA-Z0-9_%]*')

        for match in p.finditer(html):
            arg = match.group()
            arg = arg[arg.find('=')+1:]
            args.append(arg)

        try:
            url = 'http://10.243.187.170/grb/topologia_rede/www/executar_comandos_telnet_proxy.php?&arg0='+ args[0] +'&arg1='+ args[1] +'&arg2='+ args[2] +'&id_rede=15&comando='+ command_code + command
        except:
            return 0 # Hostname not valid
            sys.exit()
        #print(url)
        r = requests.post(url, auth=HTTPBasicAuth(username, password), timeout=2000)
        config = r.text
        #print(type(config))
        config = config[config.find(r"\n")+2:config.rfind(r"</pre>")]
        #print(config.find("\n")+1, config.rfind("\n"))
        #print(config)
        config = "\n".join(config.split(r"\n"))
    #elif cam.search(hostname):
    #   username = 'gdiaz'
    #    password = 'Wktkm1987#'
    #    management = 'telnet'
    elif cam.search(hostname):
        pe_username = 'gdiaz'
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            ssh.connect(hostname, username="gdiaz", password="Wktkm1987#")
        except TimeoutError:
            return -1
        chan = ssh.invoke_shell()
        buff = ''
        while buff.find('login as: ') < 0 and buff.find('Username: ') < 0 and buff.find('password: ') < 0 and buff.find('(yes/no)? ') < 0 and buff.rfind('Connection refused') < 0:
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))

        if buff.find('Username: ') != -1 or buff.find('login as: ') != -1:
            chan.send(pe_username + '\n')
            while not buff.endswith('Password: '):
                resp = chan.recv(9999)
                buff += str(resp.decode('ascii'))
            chan.send(pe_password + '\n')
        elif buff.find('password: ') != -1:
            chan.send(pe_password + '\n')
        elif buff.find('(yes/no)? ') != -1:
            chan.send("yes\n")
            while not buff.endswith('password: '):
                resp = chan.recv(9999)
                buff += str(resp.decode('ascii'))
            chan.send(pe_password + '\n')
        elif buff.find('Connection refused') == -1:
            return 0

        while not buff.endswith('#') and buff.find('RP/0/RSP0/CPU0') == -1:
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))

        chan.send('terminal length 0\n')
        chan.send('show configuration 0\n')
    else:
        #paramiko.common.logging.basicConfig(level=paramiko.common.DEBUG)
        first_octed = hostname.split('.')[0]

        if first_octed == '172':
            server_ip = '190.220.57.65'
            username = 'diego.capello'
            password = 'TZr%xbod=X!P'
            port = 6324
            management = 'telnet'
            pe_username= 'ciap'
            pe_password = "c16p_grc"
        else:
            server_ip = '200.14.205.9'
            username = 'fvrljicak'
            password = 'n4d4p4t12017'
            port = 22
            management = 'ssh'
            pe_password = "2017#T1ahPtR1"

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            ssh.connect(server_ip, username=username, password=password, port=port)
        except TimeoutError:
            return -1

        chan = ssh.invoke_shell()

        chan.send(management + ' ' + hostname + '\n')

        buff = ''
        while buff.find('Username: ') < 0 and buff.find('password: ') < 0 and buff.find('(yes/no)? ') < 0 and buff.rfind('Connection refused') < 0:
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))

        if buff.find('Username: ') != -1:
            chan.send(pe_username + '\n')
            while not buff.endswith('Password: '):
                resp = chan.recv(9999)
                buff += str(resp.decode('ascii'))
            chan.send(pe_password + '\n')
        elif buff.find('password: ') != -1:
            chan.send(pe_password + '\n')
        elif buff.find('(yes/no)? ') != -1:
            chan.send("yes\n")
            while not buff.endswith('password: '):
                resp = chan.recv(9999)
                buff += str(resp.decode('ascii'))
            chan.send(pe_password + '\n')
        elif buff.find('Connection refused') == -1:
            return 0

        buff = ''
        while not buff.endswith('#') and buff.find('RP/0/RSP0/CPU0') == -1:
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))
        chan.send("terminal length 0\n")

        if buff.find('RP/0/RSP0/CPU0') != -1:
            chan.send(command + '\n')
        else:
            chan.send(command + '\n')

        buff = ''
        time.sleep(1)
        while not buff.endswith('#'):
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))

        config = buff

        ssh.close()

    return config.split("\n") if list else config

#print(get_config('DM3000-SPOLP-DI01-PORTA28', 'show run', list=False))

def generate_ip(ip):
    """Generate second IP for /30 subnet"""
    last_octed = int(ip.split('.')[-1])
    ip_cut = ip.rfind('.') + 1
    if last_octed % 2 == 1:
        other_ip = ip[:ip_cut] + str(last_octed+1)
    else:
        other_ip = ip[:ip_cut] + str(last_octed-1)

    return other_ip

def correctSpeed(speed):
    if speed.find("K") != -1:
        return str(float(speed[:-1]) / 1024)
    return speed[:-1]

def extract_policy_speed(data, link, policy):
    policy = policy.split('_')
    if link.country.name == "BRASIL":
        if len(policy) > 5:
            data["profile"] = policy[3][2:]
            data["speed"] = correctSpeed(policy[2])
        else:
            speed = int(policy[3]) / 1024
            if speed > 1:
                speed = round(speed)
            data["speed"] = str(speed)
    if link.country.name == "MEXICO" or link.country.name == "COLOMBIA":
        if link.service == "RPV Multiservicios":
            if len(policy) == 5:
                data["profile"] = policy[2][2:]
                data["speed"] = correctSpeed(policy[3].upper())
            else:
                print(policy)
                data["profile"] = policy[-2][2:]
        else:
            if link.country.name == "COLOMBIA":
                data["speed"] = policy[0].split('-')[-1]
                if data["speed"].find("M") != -1:
                    data["speed"] = data["speed"][:-1]
                else:
                    data["speed"] = str(float(data["speed"])/1024)
            else:
                try:
                    data["speed"] = policy[-2][:-1]
                except:
                    pass
    if link.country.name == "CHILE":
        if link.service == "RPV Multiservicios":
            data["profile"] = policy[2]
            data["speed"] = policy[0][:-1]

def binary(decimal) :
    otherBase = ""
    while decimal != 0:
        otherBase  =  str(decimal % 2) + otherBase
        decimal    //=  2
    return otherBase

def convert_netmask(mask, netmask=True):
    if isinstance(mask, str) and ('\n' in mask or '\r' in mask):
        mask = mask[:-1]
    try:
        ip = IPv4Network("0.0.0.0/" + str(mask))
        if netmask:
            return ip.netmask
        return ip.prefixlen
    except Exception:
        return False

def getRoutingProtocolWithHost(data, link, hostname, country):

    if link.service == 'RPV Multiservicios':
        command = 'show ip route vrf ' + data['vrf']
    else:
        command = 'show ip route'

    output = get_config_from(country, hostname, command=command, l=False)

    if output.count('B') > 3:
        return 'B'
    return 'S'

def extract_info(config, link, hostname, country):

    data = {
        'vrf': '',
        'encap': '',
        'encapID': '',
        'profile': '',
        'client_as': '',
        'telmex_as': ''
    }

    parse = CiscoConfParse(config)

    try:
        interface = parse.find_parents_w_child("^interface", link.local_id)[0]
    except:
        print("local_id not found")
        return None

    inter_config = parse.find_all_children("^%s$" % interface)

    interface = interface.split(' ')[1]
    data['interface'] = interface

    speed_switcher = {
        'GigabitEthernet': '1000',
        'FastEthernet': '100',
        'Ethernet': '10',
        'Serial': '2',
    }

    try:
        data["speed"] = speed_switcher[interface]
    except:
        pass

    for line in inter_config:
        line = line.split(' ')
        if 'address' in line and not 'no' in line:
            if hostname.os == 'ios':
                data["pe_ip"] = line[-2]
                data["mask"] = convert_netmask(line[-1], netmask=False)
            elif hostname.os == 'xr':
                data["pe_ip"] = line[-2]
                data["mask"] = convert_netmask(line[-1], netmask=False)
            data["ce_ip"] = generate_ip(data["pe_ip"])
        elif 'vrf' in line:
            data["vrf"] = line[-1]
        elif 'ppp' in line:
            data["encap"] = 'ppp'
        elif 'frame-relay' in line:
            data["encap"] = 'frame-relay'
            data["encapID"] = line[-1]
        elif 'dot1q' in line or 'dot1Q' in line:
            if (link.local_id[0] == 'C' and link.country == 'MEXICO') or 'second-dot1q' in line:
                data["encap"] = 'dot1q'
                data["encapID"] = line[-1]
            elif link.interface == 'V.35':
                data["encap"] = 'ppp'
        elif 'hdlc' in line:
            data["encap"] = 'hdlc'
        elif 'service-policy' and 'input' in line:
            policy = line[-1]
            try:
                extract_policy_speed(data, link, policy)
            except:
                pass

    if link.country.name == "BRASIL" and hostname.os == "xr" and data['rp'] == "B":
        config += get_config_from(country, hostname.name, command="show configuration running-config router bgp 4230 vrf " + data["vrf"])
        parse = CiscoConfParse(config)

    if data.get("pe_ip", 0) and data.get("mask", 0):
        if hostname.os == "xr":
            try:
                search_string = "neighbor " + data['ce_ip']
                data["client_as"] = parse.find_all_children(search_string)[1].split(' ')[-1]
                data['rp'] = 'B'
                data['telmex_as'] = parse.find_lines('router bgp')[0].split(' ')[-1]
            except:
                data['rp'] = 'S'
        elif hostname.os == "ios":
            try:
                search_string = "neighbor " + data['ce_ip'] + " remote-as"
                data["client_as"] = parse.find_lines(search_string)[0].split(' ')[-1]
                data['telmex_as'] = parse.find_lines('router bgp')[0].split(' ')[-1]
            except:
                data['rp'] = 'S'

    return data

def placeholderReplace(filename, replacements):
    new_text = str()
    with open(filename) as file:
            for line in file:
                for string, target in replacements.items():
                    try:
                        line = line.replace(string, target)
                    except TypeError:
                        print(string, target)
                new_text = '%s%s' % (new_text, line)
    return new_text

def format_speed(speed):
    if not speed:
        return ''
    fspeed = float(speed)
    if fspeed >= 1:
        return str(int(fspeed))+"M"
    else:
        return str(int(fspeed*1024))+"K"

def create_template(link, configuration):
    perfil = str()
    policer = str()
    inter = str()
    routing = str()

    replacements = {
        '[PERFIL]': configuration.profile if configuration.profile > 10 else '0'+configuration.profile,
        '[SPEED]': format_speed(configuration.speed),
        '[KB]': str(int(float(configuration.speed)*1024)),
        '[CE]': configuration.ce_ip,
        '[PE]': configuration.pe_ip,
        '[ENCAP]': configuration.encap,
        '[ENCAPID]': configuration.encapID,
        '[C_AS]': configuration.client_as,
        '[T_AS]': configuration.telmex_as,
        '[CLIENT]': link.client.name,
        '[COUNTER]': '0',
        }

    if link.service == "RPV Multiservicios":
        print("Perfil")
        with open(os.path.join('Perfiles', '%s.txt' % replacements['[PERFIL]'])) as file:
           perfil = file.read()
        if int(configuration.profile) < 14:
            url_policer = os.path.join('Policer', 'a_policer.txt')
        else:
            url_policer = os.path.join('Policer', 'b_policer.txt')
    else:
        url_policer = os.path.join('Policer', 'policer_internet.txt')

    if configuration.interface == 'V.35' or configuration.interface == 'G703':

        num_of_inter = int(float(configuration.speed)/2)
        print(num_of_inter)
        if num_of_inter == 0:
            print("Serial less than 2mb")
            if link.service == "RPV Multiservicios":
                url_interface = os.path.join('Interface', '%s_less_than_2mb.txt' % configuration.interface)
            else:
                url_interface = os.path.join('Interface', '%s_less_than_2mb_internet.txt' % configuration.interface)
            policer = placeholderReplace(url_policer, replacements)
            inter = placeholderReplace(url_interface, replacements)
        elif num_of_inter == 1:
            print("Serial")
            if link.service == "RPV Multiservicios":
                url_interface = os.path.join('Interface', '%s.txt' % configuration.interface)
            else:
                url_interface = os.path.join('Interface', 's_internet.txt' % configuration.interface)
            inter = placeholderReplace(url_interface, replacements)
        elif num_of_inter > 1:
            print("Serial more than 2mb")
            if link.service == "RPV Multiservicios":
                url_interface = os.path.join('Interface', 'multilink.txt')
            else:
                url_interface = os.path.join('Interface', 'multilink_internet.txt')
            for counter in range(0, num_of_inter):
                replacements['[COUNTER]'] = str(counter)
                inter += placeholderReplace(os.path.join('Interface', 'multilink_%s.txt' % configuration.interface), replacements)
            inter += placeholderReplace(url_interface, replacements)
    else:
        print("Ethernet")
        if configuration.encap == "dot1q":
            print("Vlan")
            if link.service == "RPV Multiservicios":
                url_interface = os.path.join('Interface', '%s.txt' % configuration.encap)
            else:
                url_interface = os.path.join('Interface', '%s_internet.txt' % configuration.encap)

            replacements['[VLAN]'] = configuration.encapID
            policer = placeholderReplace(url_policer, replacements)
            inter = placeholderReplace(url_interface, replacements)

        else:
            print("No Vlan")
            if link.service == "RPV Multiservicios":
                print('MPLS')
                url_interface = r'%s/Interface/ethernet.txt' % (settings.BASE_DIR,)
            else:
                print('Internet')
                url_interface = r'%s/Interface/ethernet_internet.txt' % (settings.BASE_DIR,)
            policer = placeholderReplace(url_policer, replacements)
            inter = placeholderReplace(url_interface, replacements)

    if configuration.rp == 'B':
        if link.nsr[:-1] == 'B':
            if link.country == 'MEXICO':
                replacements['[CM]'] = '8151:285'
            else:
                replacements['[CM]'] = '28513:285'
            replacements['[AS_PATH]'] = ''
        else:
            if link.country == 'MEXICO':
                replacements['[CM]'] = '8151:286'
            else:
                replacements['[CM]'] = '28513:286'
            replacements['[AS_PATH]'] = 'set as-path prepend %s %s %s' % (configuration.client_as, configuration.client_as, configuration.client_as)
        routing = placeholderReplace(os.path.join('Routing', 'BGP.txt'), replacements)
    else:
        routing = placeholderReplace(os.path.join('Routing', 'default.txt'), replacements)

    return perfil + policer + inter + routing

def create_template_excel(link, template):
    #if not os.access('X:\CustomerCare\ENGINEERING SUPPORT', os.W_OK):
    #    print("No esta conectado a la VPN o no tiene accesso a la carpeta de los templates")
    #    sys.exit()

    #path = "%s\Templates\%s\%s" % (os.getcwd(), project.pgla, project.nsr)
    #if not os.path.exists(path):
    #    os.makedirs(path)
    #file = "%s\%s-%s" % (path, project.get('site_name', 0).replace("/", "-"), project.nsr)
    #template_file = open(file + ".txt", "w")
    #template_file.write(template)
    #template_file.close()

    output = BytesIO()

    workbook = xlsxwriter.Workbook(output)
    format1 = workbook.add_format({'text_wrap': True, 'bold': 1, 'border': 1, 'border_color': 'black', 'bg_color': 'blue', 'font_color': 'white', 'align': 'center'})
    format2 = workbook.add_format({'text_wrap': True, 'border': 1, 'border_color': 'black'})
    worksheet = workbook.add_worksheet('Main Circuit')

    worksheet.set_column('A:A', 70)
    worksheet.set_column('A:B', 70)

    worksheet.merge_range('A1:B1', 'Template Main Circuit', format1)

    worksheet.write(1, 1, link.site_name, format2)
    worksheet.write(2, 1, link.nsr, format2)
    worksheet.write(3, 1, link.local_id, format2)
    worksheet.write(4, 1, getProvider(link), format2)
    worksheet.write(5, 1, link.service, format2)
    worksheet.write(6, 1, link.interface, format2)
    worksheet.write(7, 1, link.config.encap, format2)
    worksheet.write(8, 1, link.config.ce_ip + '/' + link.config.mask, format2)
    worksheet.write(9, 1, link.config.pe_ip + '/' + link.config.mask, format2)
    worksheet.write(10, 1, link.config.client_as if link.config.rp == 'B' else 'None', format2)
    worksheet.write(11, 1, link.config.telmex_as if link.config.rp == 'B' else "None", format2)
    worksheet.write(12, 1, format_speed(link.config.speed), format2)
    worksheet.write(13, 1, link.config.profile, format2)
    worksheet.write(14, 1, link.config.vrf, format2)
    worksheet.write(15, 1, link.config.encapID if link.config.encapID else 'None', format2)


    worksheet.write(1, 0, 'Site Name', format1)
    worksheet.write(2, 0, 'NSR/ Circuit ID', format1)
    worksheet.write(3, 0, 'Local ID', format1)
    worksheet.write(4, 0, 'Service Provider', format1)
    worksheet.write(5, 0, 'Circuit Type', format1)
    worksheet.write(6, 0, 'Interface', format1)
    worksheet.write(7, 0, 'Encapsulation', format1)
    worksheet.write(8, 0, 'IP Wan CPE', format1)
    worksheet.write(9, 0, 'IP Wan PE', format1)
    worksheet.write(10, 0, 'Client AS', format1)
    worksheet.write(11, 0, 'Telmex AS', format1)
    worksheet.write(12, 0, 'BW', format1)
    worksheet.write(13, 0, 'QoS Profile', format1)
    worksheet.write(14, 0, 'VRF', format1)
    worksheet.write(15, 0, 'DCLI' if link.config.encap == 'frame-relay' else 'VLAN', format1)

    worksheet = workbook.add_worksheet('Main Circuit Configuration')
    worksheet.write(0, 0, 'CE Config')
    worksheet.write(1, 0, template, format2)

    workbook.close()

    output.seek(0)

    print("Done")

    return output

def getHTMLContentFromPGLA(url, method='get', username='ggutierrez', password='pgla', data=''):
    cookies = {
        "cve_paisUsuario": "USA",
        "cve_puesto": "19",
        "grant": "2%243%244%245%246%247%248%249%2410%2411%2412%2413%2414%2415%2416%2417%2418%2419%2420%2421%2422%2423%2424%2425%2426%2427%2428%2430%2431%2432%2434%2439%2440%2441%2442%2448",
        "JSESSIONID": "A80CB5F9EA210A172502FE06667B96E1",
        "uname": "GUILLERMO%2BDIAZ%2B",
        "oportunidad": "0",
        "pendientes": "28",
        "puesto": "INGENIERO",
        "usr": "GDIAZ_US"
    }
    r = getattr(requests, method)(url, auth=HTTPBasicAuth(username, password), cookies=cookies, data=data)
    html = r.text
    return html

def closeTaskcloseTask(url):

    username = 'guillelos'
    password = 'pgla'
    cookies = {
        "cve_paisUsuario": "USA",
        "cve_puesto": "19",
        "grant": "28%2432%2434%2448",
        "JSESSIONID": "D8A947A90B4318F366DC2C45294C5C40",
        "uname": "GUILLERMO%2BLOS%2B",
        "oportunidad": "43069",
        "pendientes": "27",
        "puesto": "INGENIERO",
        "usr": "GUILLELOS"
    }
    data = {
        'command': '0',
        'clave': '241474',
        'avc': '0',
        'nsr': 'MXXXN73237-0',
        'id': '8',
        'actividad': 'VALIDACION DE INVENTARIO Y GESTION',
        'finicial': '06/06/2016',
        'ffinal': '06/06/2016',
        'avance': '100',
        'slocal': '0'
    }
    r = requests.post(url, auth=HTTPBasicAuth(username, password), cookies=cookies, data=data)
    html = r.text
    print(html)

#closeTask(r"http://10.192.5.53/portalGlobal/app/nodo_plan_general_nsr_exe.jsp")

def getSiteNameFor(pgla, number):
    url = "http://10.192.5.53/portalGlobal/app/sitio_Detalle.jsp?$screen=1280&moduloCaller=VENTA%20-%20IMPLANTACION&corpEmpCve=0008863&ubicCve=" + number + "&cve=1&proyCve=" + pgla + "&nodoNo=1&cveTipoRed=1"

    html = getHTMLContentFromPGLA(url)
    html = html[html.find('<table width="95%" border="0" cellspacing="1" cellpadding="1" class="bordeTabla2">'):]
    html = html[html.find('<table width="100%" border="0" cellspacing="0" cellpadding="0">'):]
    soup = BeautifulSoup(html, 'html.parser')

    table = soup.find_all('table')[0]

    position = ''
    site_name = ''

    for table in table.children:
        for tbody in table:
            for tr in tbody:
                if isinstance(tr, NavigableString) and position == "Nombre del Sitio":
                    site_name = tr.strip()
                    if site_name:
                        position = ''
                if isinstance(tr, Tag):
                    if isinstance(tr.string, NavigableString):
                        position = tr.string.strip()

    return site_name

def getSpeedInterfaceProfile(pgla, addressNumber, nsr):

    speed = None
    inter_type = None
    profile_number = None
    data_alta = []
    data_baja = {}
    site_name = ''
    url = "http://10.192.5.53/portalGlobal/app/nodo_Detalle.jsp?$screen=1280&moduloCaller=VENTA%20-%20IMPLANTACION&proyCve="+ pgla +"&cve="+ addressNumber

    html = getHTMLContentFromPGLA(url)

    p = re.compile("fnSitioDetalleNodo\(CVE_TIPO_RED\,\s\'\d{5,8}\',\'" + addressNumber + "\',\'VENTA\s-\sIMPLANTACION\',\s\'\d{7,9}\',\s\'\d{5,8}\',\s\'1\'\s+\)")
    m = p.search(html)

    if m:
        p1 = re.compile("\d{5,8}")
        for i, match in enumerate(p1.finditer(m.group())):
            if i == 2:
                site_name = getSiteNameFor(pgla, match.group())

    soup = BeautifulSoup(html, 'html.parser')

    table = safe_list_get(soup.find_all('table'), -4, 0)

    nsrRE = re.compile(nsr)
    puerto = re.compile("PUERTO\s\d+")
    interfaz = re.compile("ULTIMA MILLA,")
    profile = re.compile("PERFIL\s\d+")
    alta = re.compile("ALTA|PRODUCTO\sACTIVO")
    baja = re.compile("BAJA DE PRODUCTO")
    speed_count = 0
    links_number = 0
    count = 0
    nsr_not_correct = True
    for tr in table.children:
        for td in tr:
            if isinstance(td, Tag):
                #print(td.string)
                try:
                    info = str(td.string)
                    #print(info)
                except:
                    continue
                #print(info.encode('utf-8'))
                if puerto.search(info):
                    #print(info)
                    m = puerto.search(info)
                    speed = info[m.start():].split(" ")[1]

                    if speed.find("K") != -1:
                        speed = speed.replace(',', '')
                        speed = float(speed[:speed.find("K")]) / 1024
                        if speed > 1:
                            speed = str(int(speed))
                        else:
                            speed = str(speed)
                    #elif "Gbps" in unit and speed.isdecimal():
                    #    speed = str(int(speed) * 1024)

                elif interfaz.search(info):
                    if info.rfind('Fast Ethernet/') != -1:
                        #print('Fast')
                        inter_type = 'Ethernet'
                    elif info.rfind('Electrico') != -1:
                        #print('Electrico')
                        inter_type = 'Ethernet'
                    elif info.rfind('G.703') != -1 or info.rfind('G703') != -1:
                        #print("G703")
                        inter_type = 'G703'
                    elif info.rfind('V.35') != -1:
                        inter_type = 'V.35'
                    elif info.rfind('Optic') != -1:
                        #print('BNC')
                        inter_type = info.split(" ")[-3]
                    else:
                        inter_type = info.split(" ")[-1]
                    if inter_type.find("/") != -1:
                        inter_type = inter_type[:inter_type.find("/")].upper()
                elif profile.search(info):
                    if info.rfind("FIJO") == -1:
                        profile_number = info.split(" ")[-1][1:-1]
                    else:
                        profile_number = info.split(" ")[-2][1:-1]
                    #print(pgla, nsr, profile_number)
                    try:
                        if int(profile_number) < 10:
                            profile_number = "0" + profile_number
                    except:
                        pass
                elif baja.search(info):
                    if speed:
                        data_baja['speed'] = speed
                        speed = None
                    elif inter_type:
                        data_baja['interface'] = inter_type
                        inter_type = None
                    elif profile_number:
                        data_baja['profile'] = profile_number
                        profile_number = None
                elif nsrRE.search(info):
                    nsr_not_correct = False
                    #print('NSR correct')
                    count = 3
                elif alta.search(info):
                    if nsr_not_correct:
                        #print('ALTA not correct')
                        break
                    #print('ALTA correct')
                    if safe_list_get(data_alta, links_number, 0) == 0:
                        data_alta.append({'nsr': nsr})
                    if speed:
                        #print(speed)
                        if speed_count == 1:
                            links_number += 1
                        if safe_list_get(data_alta, links_number, 0) == 0:
                            data_alta.append({'nsr': nsr})
                        data_alta[links_number]['speed'] = speed
                        speed = None
                        speed_count += 1
                    elif inter_type:
                        #print(inter_type)
                        data_alta[links_number]['interface'] = inter_type
                        inter_type = None
                    elif profile_number:
                        #print(profile_number)
                        data_alta[links_number]['profile'] = profile_number
                        profile_number = None
                if not nsrRE.search(info) and not nsr_not_correct:
                    count -= 1
                    #print('count ', count)
                    if count == 0:
                        nsr_not_correct = True
                        #print('NSR not correct')
    #print(data_alta)
    return data_alta, data_baja, site_name

#print(getSpeedInterfaceProfile('43475', '1'))

def getAddressSpeedInterfaceProfileFromPGLA(pgla, nsr):

    data = {}

    url = "http://10.192.5.53/portalGlobal/app/proy_lista_circuito_nodo.jsp?$screen=1280&proyCve=" + pgla + "&moduloCaller=VENTA%20-%20IMPLANTACION"

    html = getHTMLContentFromPGLA(url)

    html = html[html.find('<table width="897" border="0" cellpadding="1" cellspacing="1" class="bordeTabla3">'):]

    soup = BeautifulSoup(html, 'html.parser')

    nsrRE = re.compile(nsr)
    address = re.compile("\d+(.*?)\-(.*?)")

    count = 0
    for child in soup.table.contents:
        if nsrRE.search(str(child)):
            count += 1
        elif count == 1:
            count += 1
        elif count == 2:
            child = str(child)
            match = address.search(child)

            if match:
                number = "".join([s for s in match.group() if s.isdigit()])
                alta, baja, site_name = getSpeedInterfaceProfile(pgla, number, nsr)

                child = child[match.end():child.find("</div>")]
                child = child.replace("<br>", " ")
                child = child.replace("\xa0", " ")
                child = child.replace("<br/>", " ")
                child = child.replace("\n", " ")
                child = child.replace('<div class="textoDomicilio">', " ")

                data['address'] = child.strip()
                data['links'] = alta
                data['site_name'] = site_name
                data['circuit_id'] = get_circuit_id(pgla, nsr)
                return data

    return False

def getParticipansWithPGLA(pgla):
    from django.contrib.auth.models import User
    url = 'http://10.192.5.53/portalGlobal/app/proy_Participantes.jsp?$screen=1280&cve=' + pgla
    html = getHTMLContentFromPGLA(url)
    html = html[html.find('<table width="95%" border="0" cellspacing="1" cellpadding="1" class="bordeTabla2">'):]
    html = html[html.find('<table width="100%" border="0" cellspacing="0" cellpadding="0">'):]
    soup = BeautifulSoup(html, 'html.parser')

    table = soup.find_all('table')[0]

    participans = []
    position = ''

    for table in table.children:
        for tbody in table:
            for tr in tbody:
                if isinstance(tr, Tag):
                    if isinstance(tr.string, NavigableString):
                        position = tr.string[4:-5]
                if isinstance(tr, NavigableString):
                    member_info = tr.string.split("-")
                    if len(position) > 0 and len(member_info) > 1:
                        first_name, last_name = member_info[0].lower().strip().split(' ', 1)
                        username = first_name+last_name.replace(' ', '')
                        user, created = User.objects.get_or_create(username=username)
                        #print(user.username)
                        if not created:
                            participans.append(user)
                            continue
                        user.first_name = first_name.capitalize()
                        user.last_name = last_name.capitalize()
                        user.profile.company = member_info[1].strip()
                        user.profile.position = position
                        if len(member_info) == 4:
                            user.email = member_info[3].strip()
                            user.profile.number = member_info[2].strip()
                        elif len(member_info) == 6:
                            user.email = member_info[5].strip()
                            user.profile.number = member_info[2].strip() + "-" + member_info[3].strip() + "-" + member_info[4].strip()
                        else:
                            user.email = member_info[4].strip()
                            user.profile.number = member_info[2].strip() + "-" + member_info[3].strip()

                        user.save()
                        participans.append(user)

    return participans

#getParticipansWithPGLA(40274)

def fixLocalID(localID_list, speed):
    primary = ''
    secondary = ''
    new_localID_list = []

    numberOfInt = speed/2

    for i, local_id in enumerate(localID_list):
        if i < numberOfInt:
            primary += ' ' + local_id
        else:
            secondary += ' ' + local_id

    if primary != '':
        new_localID_list.append(primary.strip())
    if secondary != '':
        new_localID_list.append(secondary.strip())

    return new_localID_list

def getDataFromPGLA(nsr=None):
    url = "http://10.192.5.53/portalGlobal/reportes/reporteEjCIAPDetalle.jsp?tipoReporte=1&estatusserv=ENPROCESO&tiposerv=&fechaInicioPen=&fechaFinPen=&fcambioestatus=&cliente=&nombreCAPL=&nombrePM=&nombreIMP=&nombreIS=&estatus="

    keys = [
        'number', 'client', 'client_segment',  'pm', 'imp', 'is', 'capl', 'pgla', 'nsr', 'local_id', 'service', 'movement',
        'state', 'motive', 'country_a', 'country_b', 'duedate_ciap', 'duedate_acc', 'entraga_ciap',
        'loop-ready', 'recepcion_ciap', 'billing_date', 'cnr', 'ddf', 'daf', 'observation', 'duration',
        'duracion_contract',
    ]

    collection = []

    html = getHTMLContentFromPGLA(url)
    html = html[html.find('<table width="100%" ID="reporte" border="0" cellpadding="1" cellspacing="1" class="bordeTabla">'):]
    soup = BeautifulSoup(html, 'html.parser')

    #nsr = re.compile(nsr)

    for table in soup.table.contents:
        for tbody in table:
            #tr_count = 0
            for tr in tbody:
                #if nsr.search(str(tr)):
                    td_count = 0
                    document = {}
                    for td in tr:
                        if isinstance(td, Tag) and len(keys) > td_count:
                            #elif keys[td_count] == 'local_id' and not td.string:
                            #    document[td_count] == get_info_for_site_name(document['nsr'], 'local_id')
                            if keys[td_count] == 'country_a' and document['local_id']:
                                regex = {
                                    'MEXICO': "(A|C|D|F|S)([B|T]\d{1}|\d{2})-\d{4}-\d{4}",
                                    'BRASIL': "(CSL|CEM|SBS|MBB|LDA|MGA|SJC|SOO|GNA|SGO|PAS|PGE|RPO|ULA|GRS|SPO|PAE|CAS|RJO|SJP|SDR|BPI|ITU|CTA|MNS|SOC|CBO|MCO|UPS|FLA|JBO|SRR|PSO|RCO|LNS|VGA|ABS|CSC|YCA|ETA|JVE|LNA|BRU|JAI|SPA|AMA|STB|NTL|MCL|CBM|NHO|PLT|BRE|LGS|GAMA|PTA|FSA|PTS|SBO|CE|FLA|PWM)\/(IP|MULTI|FAST)\/(\d{5})",
                                    'PERU': "\d{7}",
                                    'HONDURAS': "\d{6}",
                                    'GUATEMALA': "\d{8}",
                                    'CHILE': "\d{2}-\d{2}-\d{10}",
                                    'COLOMBIA': "^\w{3}\d{4}$",
                                    'EL SALVADOR': "IP\d{7}",
                                    'ESTADOS UNIDOS': 'ITFS\-\d{11}',
                                    'ARGENTINA': '\d{7}',
                                    'VENEZUELA': '',
                                    'ECUADOR': "asd",
                                    'AUSTRIA': "",
                                    'PARAGUAY': "",
                                    'NICARAGUA': "",
                                    'ESPAÑA': "",
                                    "URUGUAY": "",
                                    'BOLIVIA': "",
                                    'REPUBLICA DOMINICANA': "^\d{3}-\d{3}-\d{4}$",
                                }

                                p = re.compile(regex[td.string])
                                document['local_id'] = [match.group() for match in p.finditer(document['local_id'])]

                            document[keys[td_count]] = td.string

                            if keys[td_count] in ['pm', 'imp', 'is', 'capl']:
                                pm = ' '.join(td.string.split())
                                document[keys[td_count]] = pm

                            #print(document)
                            td_count += 1

                    if document != {} and document['imp'] in imp_list:
                        if nsr:
                            if document['nsr'] == nsr:
                                collection.append(document)
                                break
                        else:
                            collection.append(document)

                    #tr_count += 1

    bar = progressbar.ProgressBar()
    for doc in bar(collection):
        asip = getAddressSpeedInterfaceProfileFromPGLA(doc['pgla'], doc['nsr'])
        #print(asip)
        doc.update(asip)
        if doc['local_id'] and doc['alta'].get('interface', '') in ['G.703', 'V.35'] and float(doc['alta'].get('speed', 0)) > 2:
                #print(doc['nsr'], doc['alta'])
                doc['local_id'] = fixLocalID(doc['local_id'], int(float(doc['alta']['speed'])))

        participans = getParticipansWithPGLA(doc['pgla'])
        doc['participans'] = participans

    return collection if len(collection) > 1 else safe_list_get(collection, 0, 0)

#print(getDataFromPGLA('MXXXM0092233-6'))

def createEXCEL(output, header, information, data):
    from django.utils.encoding import force_text
    #directory = filepath[:filepath.rfind("\\")]

    #if not os.path.exists(directory):
    #    os.makedirs(directory)

    workbook = xlsxwriter.Workbook(output, {'strings_to_urls': False})
    format1 = workbook.add_format({'text_wrap': True, 'bold': 1, 'border': 1, 'border_color': 'black', 'bg_color': 'blue', 'font_color': 'white', 'align': 'center'})
    format2 = workbook.add_format({'text_wrap': True, 'border': 1, 'border_color': 'black'})

    worksheet = workbook.add_worksheet("RFS")

    worksheet.set_column('A:A', 70)
    worksheet.set_column('A:B', 70)

    worksheet.merge_range('A1:B1', header, format1)

    for row, value in enumerate(information):
        worksheet.write(row+1, 0, value, format1)

    for row, value in enumerate(data):
        if type(value) != unicode or type(value) != str:
            worksheet.write(row + 1, 1, value, format2)
        else:
            worksheet.write(row+1, 1, value.encode('utf-8'), format2)

    workbook.close()

    output.seek(0)
    return output

def createRFS(document):
    data = {}
    #filepath = "%s\Templates\%s\%s\PGLA-%s-%s-%s.xlsx" % (os.getcwd(), data['pgla'], data['nsr'],
    #                                             data['pgla'], data['nsr'], data['movement'])
    output = BytesIO()
    #print(filepath)
    header = "INFORMACION PARA PROGRAMACION DE ACTIVIDAD IMPLEMENTACIONES"

    if not document.movement == 'BAJA':
        data['local_id'] = ''
        data['hostname'] = ''
        data['pe_ip'] = ''
        data['ce_ip'] = ''
        data['vrf'] = ''
        data['scheme'] = ''
        for link_number, circuit in enumerate(document.alta):
            if safe_list_get(document.links, link_number, 0):
                print(link_number, document.links[link_number])
                data['local_id'] += circuit['local_id'] + ' '
                data['hostname'] += document.links[link_number]['hostname'] + ' '
                data['pe_ip'] += document.links[link_number]['pe_ip'] + ' '
                data['ce_ip'] += document.links[link_number]['ce_ip'] + ' '
                data['vrf'] += document.links[link_number]['vrf'] + ' '

        if document.client == "PEPSICO INC":
            tier = get_info_for_site_name(document.nsr, 'tier')
            metal = get_info_for_site_name(document.nsr, 'metal')
            data['sla'] = 'Tier %s (%s)' % (tier, metal)
            if len(document.alta) >= 2:
                data['scheme'] += 'Primary/Secondary'
            else:
                data['scheme'] += 'Primary'
        else:
            if len(document.alta) >= 2:
                data['scheme'] += 'Primary/Secondary'
                data['sla'] = '99.80%'
            else:
                data['scheme'] += 'Primary'
                data['sla'] = '99.50%'

    #print(data)

    if document.movement == "ALTA":

        information = [u"Tipo de cambio:", u"Fecha y hora de preactivacion:", u"Cambio Uninet Aprobado/PGLA:",
                       u"No. De Cambio", u"Fecha y Hora entrega de implementacion:", u"Supervisor/Analista CNOC:",
                       u"Administrador de proyecto:", u"Supervisor de Ingenieria:", u"CLIENTE:", u"Nombre del Sitio:",
                       u"Domicilio Fisico:", u"Pais en el que se entrega el servicio:", u"SID:", u"Numero Sucursal /Tienda:",
                       u"Referencia (NSR): ", u"ID LOCAL:", u"Proveedor del Servicio", u"IP Wan PE:", "IP Wan Cliente:",
                       u"POP Uninet:", u"Ancho de banda:", u"Tipo de Servicio:", u"Esquema:", u"Equipo CPE (modelo):", u"Dueño de CPE:",
                       u"IP de Administracion (CPE):", u"SLA (Disponibilidad):", u"Version QoS:", u"Perfil QoS:",
                       u"default (Version 2.5)", u"datos_criticos (Version 2.5)", u"voz (Version 2.5)", u"voz (Version 2.5)",
                       u"Contacto Cliente", u"Teléfono contacto", u"Contacto de cliente que valido activacion de servicio:",
                       u"VRF cliente:", u"IP LAN CLIENTE:", u"PROTOCOLO IGP:", u"Observaciones:"]
        data = [document.movement, document.loop_ready, document.pgla, "", document.loop_ready, "", document.pm,
                "MAURICIO DAVALOS", document.client, "", document.address, document.country_a, document.site_name, "", document.nsr,
                data['local_id'], getProvider(document), data['pe_ip'], data['ce_ip'], data['hostname'], document.alta[0].get('speed')+"Mbps" if document.alta[0].get('speed') else document.alta[0].get('speed'),
                document.service, data['scheme'], "", "", "", data['sla'], "3", document.alta[0].get('profile', 'None'), "", "", "", "", "",
                "", "", data['vrf'], "", "", "Sin Observaciones"]

        createEXCEL(output, header, information, data)
    elif document.movement == "CAMBIO":
        information = [u"Tipo de cambio:", u"Fecha y hora de preactivacion:", u"Cambio Uninet Aprobado/PGLA:",
                       u"No. De Cambio", u"Fecha y Hora entrega de implementacion:", u"Supervisor/Analista CNOC:",
                       u"Administrador de proyecto:", u"Supervisor de Ingenieria:", u"CLIENTE:", u"Nombre del Sitio:",
                       u"SID:", u"Referencia (NSR Actual): ", u"ID LOCAL (Actual):", u"Proveedor del Servicio (Actual)",
                       u"Referencia (NSR Nueva): ", u"ID LOCAL (Nuevo):", u"Proveedor del Servicio (Nuevo)",
                       "IP Wan Cliente (Actual):", "IP Wan Cliente (Nueva):",
                       u"POP Uninet:", u"Ancho de banda (Actual):", u"Ancho de banda (Nuevo):", u"SLA (Disponibilidad):", u"Version QoS:", u"Perfil QoS:",
                       u"default (Version 2.5)", u"datos_criticos (Version 2.5)", "voz (Version 2.5)",
                       u"Contacto Uninet", "Observaciones:"]
        data = [document.movement, document.loop_ready, document.pgla, "", document.loop_ready, "", document.pm, "MAURICIO DAVALOS", document.client, "",
                document.site_name, document.nsr, data['local_id'], getProvider(document), document.nsr,     data['local_id'], getProvider(document), "",data['ce_ip'], data['hostname'],
                format_speed(document.baja.get('speed', 0)), format_speed(document.alta[0].get('speed', 0)), data['sla'], "3", document.alta[0].get('profile', 0), "", "", "", "", "Sin Observaciones"]
        createEXCEL(output, header, information, data)
    elif document.movement == "BAJA":
        information = [u"Tipo de cambio:", u"Cambio Uninet Aprobado/PGLA:",
                       u"No. De Cambio", u"FecreateEXCELcha y Hora baja de servicio:", u"Supervisor/Analista CNOC:",
                       u"Administrador de proyecto:", u"Supervisor de Ingenieria:", u"CLIENTE:", u"Pais", u"SID:",
                       u"Referencia (NSR): ", u"ID LOCAL:", u"Proveedor del Servicio", u"Domicilio Fisico",
                       u"Contacto Uninet", u"Observaciones:"]
        data = [document.movement, document.pgla, '', document.loop_ready, '', document.pm, "MAURICIO DAVALOS", document.client, document.country_a,
                '', document.nsr, data['local_id'], getProvider(document), document.address, '', "Sin Observaciones"]
        output = createEXCEL(output, header, information, data)

    return output

def returnBrasilList(centroOrLATAM, regionOrPE):
    url = "http://10.243.187.170/grb/topologia_rede/www/consulta_topologia.php?nome_mapa=%s+%s&PERMS1=Mjs5OzczOzc0Ozc1Ow==&remote_user=dcapello@latam&tcos=1462975961028&sintomas=19;1;2;3;4;5;6;26;28;8;9;10;29&al_ral_fech=N"
    r = requests.get(url % (centroOrLATAM, regionOrPE))
    html = r.text
    html = html[html.rfind('pos_vertex_top.value'):html.rfind(';parent.posiciona')]

    return html

def downloadBrasilConfigs():

    start = datetime.now()
    html = returnBrasilList('centros', 'latam')
    p = re.compile("[A-Z]+")

    for match_region in p.finditer(html):
        region = match_region.group()
        print('Downloading configs for '+region)
        html = returnBrasilList('LATAM', region)

        p = re.compile("[A-Z]+[0-9]+.[A-Z]+")
        for match_pe in p.finditer(html):
            pe = match_pe.group()
            config = get_config(pe, 'show running-config', list=False)
            path = '%s\Configuration\Brasil\%s' % (os.getcwd(), region)
            if not os.path.exists(path):
                os.makedirs(path)
            with open('%s\%s.txt' % (path, pe), 'w') as file:
                file.write(config)
            print(pe+"'s Configuration was Saved")
    end = datetime.now()

    totalMin = (end - start).total_seconds()/60
    print(totalMin)

def downloadMexicioConfigs():
    config = get_config('vpn-mex-popocatepetl-51', 'show bgp vrf grey_mgmt_vpn_Uninet')
    print(config)

server_credencials = {
    '200.14.205.9': ('fvrljicak', 'T3lm3x!', '10.10.66.1'),
    '190.220.57.65': ('diego.capello', 'TZr%xbod=X!P', '172.31.128.31')
}

def downloadConfigsUsingSSH(ip):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    #Validate the correct IP Address
    if not server_credencials.get(ip):
        return 0

    username, password, router = server_credencials.get(ip)

    try:
        ssh.connect(ip, username=username, password=password)
    except TimeoutError:
        return -1

    chan = ssh.invoke_shell()

    # Ssh and wait for the password prompt.
    chan.send('./CHICONORTE2.sh\n')
    buff = ''
    print('Connected to CHICONORTE2')
    while not buff.endswith('CHICONORTE2#'):
        resp = chan.recv(9999)
        print(resp)
        buff += str(resp.decode('ascii'))
    print('Prompt ready for CHICONORTE2')
    # Execute whatever command and wait for a prompt again.
    chan.send('terminal length 0\n')
    chan.send('show ip route ospf | inc /32\n')
    buff = ''
    while not buff.endswith('CHICONORTE2#'):
        resp = chan.recv(9999)
        buff += str(resp.decode('ascii'))
    print('Commands "terminal length 0" and "show ip route ospf | inc /32" executed')
    # Now buff has the data I need.
    p = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/32")
    #print(buff)
    chan.send('exit\n')
    print("Exiting from CHICONORTE2")
    while not buff.endswith('[fvrljicak@hendrix ~]$ '):
        resp = chan.recv(9999)
        buff += str(resp.decode('ascii'))
    for match_pe in p.finditer(buff):
        pe = match_pe.group()[:-3]
        print(pe)
        if pe == '10.160.31.145' or pe == '172.22.115.108' or pe == '201.125.254.0' or pe == '201.125.255.0':
            continue
        #print(buff)
        chan.send('ssh %s\n' % (pe,))
        buff = ''
        while buff.find('password: ') < 0 and buff.find('(yes/no)? ') < 0 and buff.find('Connection refused') < 0:
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))

        if buff.find('password: ') != -1:
            chan.send("AYSKn1'H=D\n")
        elif buff.find('(yes/no)? ') != -1:
            chan.send("yes\n")
            while not buff.endswith('password: '):
                resp = chan.recv(9999)
                buff += str(resp.decode('ascii'))
            chan.send("AYSKn1'H=D\n")
        else:
            continue
        buff = ''
        print("Connected to PE " + pe)
        while not buff.endswith('#') and buff.find('RP/0/RSP0/CPU0') == -1:
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))

        chan.send("terminal length 0\n")

        if buff.find('RP/0/RSP0/CPU0') != -1:
            chan.send("show configuration run\n")
        else:
            chan.send("show run\n")
        print("Ran commnads 'terminal length 0' and 'show run'")
        buff = ''
        while not buff.endswith('#'):
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))
        print(7)
        if buff.find('Command authorization failed') < 0 and len(buff) > 500:
            path = '%s\Configuration\Colombia' % (os.getcwd())
            if not os.path.exists(path):
                os.makedirs(path)
            with open('%s\%s.txt' % (path, pe), 'w') as file:
                file.write(buff)
        chan.send('exit\n')
        buff = ''
        while not buff.endswith('[fvrljicak@hendrix ~]$ '):
            resp = chan.recv(9999)
            buff += str(resp.decode('ascii'))

    ssh.close()

def updateCreateCommandList(router, command):
    print(command)
    sub_command_list = get_config(router, command + " ?")
    print(sub_command_list)
    for sub_command in sub_command_list:
        #print(sub_command)
        if '<cr>' in sub_command or 'WORD' in sub_command or 'Unrecognized' in sub_command or 'A.B.C.D' in sub_command or '#' in sub_command:
            print("No more child commands")
            continue
        updateCreateCommandList(router, "%s %s ?" % (command, sub_command.split(' ')[2]))

#updateCreateCommandList('IACC01.NTL', 'traceroute')

def readPDF():
    from pdfminer.pdfparser import PDFParser, PDFDocument

    # Open a PDF document.
    fp = open(r"C:\Users\Guillermo.Diaz\PycharmProjects\TemplateGenerator\Templates\43755\PEXXM0101526-2\5910.pdf", 'rb')
    parser = PDFParser(fp)
    document = PDFDocument(parser)
    print(document.is_extractable)
    # Get the outlines of the document.
    outlines = document.get_outlines()
    for (level,title,dest,a,se) in outlines:
        print (level, title)
#readPDF()

def readingPDF():
    import PyPDF2
    pdfFileObj = open(r"C:\Users\Guillermo.Diaz\PycharmProjects\TemplateGenerator\Templates\43755\PEXXM0101526-2\5910.pdf", 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj, strict=False)

    for pageNum in range(pdfReader.numPages):
        pageObj = pdfReader.getPage(pageNum)
        print(pageObj.extractText())

#readingPDF()

def runSSHCommand(host, command):
    import subprocess
    import sys

    ssh = subprocess.Popen([r"C:\Users\Guillermo.Diaz\Downloads\putty.exe", "%s" % host, command],
                           shell=False,
                           stdout=subprocess.PIPE,
                           stderr=subprocess.PIPE)
    result = ssh.stdout.readlines()
    if result == []:
        error = ssh.stderr.readlines()
        print(sys.stderr, "ERROR: %s" % error)
    else:
        print(result)

#runSSHCommand('190.220.57.65', 'ls')

class link_data:
    pe_ip = "10.251.128.1"
    ce_ip = "10.251.128.2"
    link_type = 'Primary'
    hostname = "vpn-son-obregon-16"
    profile = "22"
    vrf = "VRFNAME"
    speed = "8MB"

def emailGenerator(data, attachment, rfs=True):

    email = ''

    for participan in data['participans']:
        if participan['name'] == data['pm']:
            email = participan['email']

    if rfs:
        subject = 'RFS %s - PGLA: %s - NSR: %s - REQUEST: %s - TYPE: %s' % \
                  (data['client'], data['pgla'], data['nsr'], data['movement'], data['service'])
    else:
        subject = '%s (PGLA: %s | NSR: %s) Template' % (data.get('site_name', 0), data['pgla'], data['nsr'])

    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    msg = outlook.OpenSharedItem("%s\RFS.msg" % (os.getcwd()))
    msg.Attachments.Add(attachment)
    msg.To = msg.To + ";" + email #pm_email[data['pm']]
    msg.Subject = subject
    msg.Display(True)
    #msg.Send()



def readEmails():
    import win32com.client
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)
    messages = inbox.Items
    for message in messages:
        print(message.Body)

def validatingConfig(data, dataInEquipment):
    if data['alta']['speed'] != dataInEquipment['speed']:
        print('Wrong speed configured: PGLA %s, PE: %s' % (data['alta']['speed'], dataInEquipment['speed']))
    if data['alta']['interface'] != dataInEquipment['interface']:
        print('Wrong interface configured: PGLA %s, PE: %s' % (data['alta']['interface'], dataInEquipment['interface']))
    if data['alta']['profile'] != dataInEquipment['profile']:
        print('Wrong interface configured: PGLA %s, PE: %s' % (data['alta']['profile'], dataInEquipment['profile']))


def updateDatabase(collection, nsr=None):

    dataPGLA = getDataFromPGLA(nsr)

    if nsr and dataPGLA:
        collection.update({'nsr': nsr}, {"$set": dataPGLA}, upsert=True)
    else:
        for documentPGLA in dataPGLA:
            document = collection.find_one({'nsr': documentPGLA["nsr"], 'complete': {'$exists': False}})
            if document and document['local_id']:
                documentPGLA.pop('local_id')
            collection.update({'nsr': documentPGLA["nsr"]}, {"$set": documentPGLA}, upsert=True)

    #print(collection.find_one({"nsr": nsr}))

#updateDatabase('BRXXN0096588-1')

link_number_to_string = {
        0: 'primary',
        1: 'secondary',
        2: 'tertiary',
        3: 'quaternary',
    }

def downloadConfiguration(collection, document):
    add = {'links': []}
    if isinstance(document['local_id'], list) and document.get('links', 0):
        for link_number, local_id in enumerate(document['local_id']):
            if safe_list_get(document['links'], link_number, 0) and document['links'][link_number].get('hostname', 0):
                print(document['links'][link_number]['hostname'])
                if document['country_a'] == 'MEXICO':
                    config = get_config(document['links'][link_number]['hostname'], "show configuration", list=False)
                else:
                    config = get_config(document['links'][link_number]['hostname'], "show version | inc Version", list=False)

                if config == CONNECTION_PROBLEM:
                    print("You are not connected to the VPN")
                    print("Please connect to the VPN and try again")
                elif config == INVALID_HOSTNAME:
                    print("The hostname you entered is not valid")
                    print("Please verify the hostname and try again")
                else:
                    if is_xr(config):
                        config = get_config(document['links'][link_number]['hostname'], "show configuration running-config")
                    else:
                        config = get_config(document['links'][link_number]['hostname'], "show running-config")

                    link_config = extract_info(config, document, link_number)

                    if not link_config:
                        print('Cant find configuration for %s in %s' % (local_id, document['links'][link_number]['hostname']))
                        continue
                    document['links'][link_number].update(link_config)
                    print(document['links'][link_number])
                    collection.update({'nsr': document["nsr"]}, {"$set": {'links': document['links']}}, upsert=True)
            else:
                print("Link %s has no hostname configured" % (link_number_to_string[link_number]))

def peLookupWith(local_id):
    url = 'http://10.243.187.170/grb/topologia_rede/www/consulta_busca_avancada_jsrs.php?metodo=get_designacao&rede=15&designacao=' + local_id + '&componentes_multilink=0&tcos=1480431069052%20HTTP/1.1'
    username = 'dcapello@latam'
    password = 'latam'

    p = re.compile("(GACC|IACC|UACC|NGACC|TACC|NIACC|QGACC|ACC|XACC)\d{2}.(AJU|BHE|BLM|BRE|BRU|BSA|BVA|CAS|CBA|CEM|CIM|CPE|CSL|CTAMC|CTA|FLA|FNS|FSA|GNA|GVS|JFA|JPA|LDA|MBA|MCO|MGA|MNS|MPA|NTL|OCO|PAE|PAEMF|PTA|PVO|RBO|RBO|RCE|RJOAC|RJOEN|RJOST|RJO|RPO|SBO|SDR|SJC|SLS|SNE|SOC|SPOLP|SPOMB|SPOPH|SPO|SRM|STS|TSA|ULA|VTA)")

    try:
        response = requests.post(url, auth=HTTPBasicAuth(username, password), timeout=2000)
    except requests.exceptions.ConnectionError as e:
        return 0 # No connection to Looking Glass

    html = response.text
    m = p.search(html)

    if m:
        return m.group()

    return 0

def getOtherPE(country, local_id):
    hostname = Hostname.ob
    hostname = None
    client = MongoClient()
    db = client.info
    doc = db.interfaces.find_one({'DESCRIPTION': re.compile('.*' + local_id + '.*')})

    if doc:
        if country == 'COLOMBIA':
            hostname = colo_host_to_ip[doc['ROUTER']]
        else:
            hostname = doc['ROUTER']
    print(hostname, "aqui esta")
    return hostname

def get_circuit_id(pgla, nsr):
    url = "http://10.192.5.53/portalGlobal/app/proy_lista_circuito_nodo.jsp?$screen=1280&proyCve=" + pgla + "&moduloCaller=VENTA%20-%20IMPLANTACION"
    html = getHTMLContentFromPGLA(url)
    html = html[html.find('<table width="897" border="0" cellpadding="1" cellspacing="1" class="bordeTabla3">'):]
    soup = BeautifulSoup(html, 'html.parser')

    table = soup.find_all('table')[0]

    for table in table.children:
        for tbody in table:
            for tr in tbody:
                if isinstance(tr, Tag):
                    tr = str(tr)
                    if nsr in tr:
                        match = re.search("div_\d{6}", tr)
                        if match:
                            return match.group().split("_")[1]
                        else:
                            return

def totalTimeSpan(pgla, nsr):
    total_days = 0
    onHoldDate = 0

    data_list = downloadOnHoldJSP(pgla, nsr)

    for date, state in data_list:
        if state == 'INSTALACION SUSPENDIDA':
            onHoldDate = date
        elif onHoldDate and state == 0:
            days = (date - onHoldDate).days
            total_days = total_days + days
            onHoldDate = 0

    if data_list[-1][-1] == 'INSTALACION SUSPENDIDA':
        days = (datetime.now() - data_list[-1][0]).days
        total_days = total_days + days

    return total_days

def downloadOnHoldJSP(pgla, nsr):

    circuit = get_circuit_id(pgla, nsr)

    if not circuit:
        return [(0, 0)]

    url = "http://10.192.5.53/portalGlobal/reportes/reporteHistoricoEstatus.jsp?$screen=1366&cveOp=" + pgla + "&circuito=" + circuit + "&nsr=" + nsr
    # print(url)
    html = getHTMLContentFromPGLA(url)
    if html.find('this.close()') != -1:
        print('No estan las fechas en el PGLA')
        return [(0, 0)]
    soup = BeautifulSoup(html, 'html.parser')
    date_list = []

    table = soup.find_all('table')[3]

    for table in table.children:
        for tbody in table:
            for tr in tbody:
                # print(tr)
                match = re.match('\d{2}\/\d{2}\/\d{4}', tr)
                if match:
                    date = match.group()
                    date_list.append([datetime.strptime(date, '%d/%m/%Y'), 0])
                else:
                    match = re.match('INSTALACION SUSPENDIDA', tr)
                    if match:
                        date_list[-1][-1] = 'INSTALACION SUSPENDIDA'
    # print(date_list)
    return date_list

def suspender(link):
    date = datetime.now()

    if not link.circuit_id:
        link.circuit_id = get_circuit_id(str(link.pgla), link.nsr)
        link.save()

    url = "http://10.192.5.53/portalGlobal/app/proy_suspender.jsp"

    data = {
        'lista': link.circuit_id,
        'proyCve': link.pgla,
        'fecha': date.strftime("%d-%m-%Y %H:%M:%S"),
        'motivo': '4',
        'observaciones': 'suspended',
        'mes': datetime.now().strftime("%m"),
        'dia': datetime.now().strftime("%d"),
        'cveoportunidad': link.pgla,
        'command': '1',
        'anio': datetime.now().strftime("%Y")
    }
    data = getHTMLContentFromPGLA(url, method='post', username='ldiaz', password='12345', data=data)

def reactivar(link):
    date = datetime.now()

    if not link.circuit_id:
        link.circuit_id = get_circuit_id(str(link.pgla), link.nsr)
        link.save()

    url = "http://10.192.5.53/portalGlobal/app/proy_reactivar.jsp"

    data = {
        'lista': link.circuit_id,
        'proyCve': link.pgla,
        'dia': datetime.now().strftime("%d"),
        'dias': '1',
        'mes': datetime.now().strftime("%m"),
        'anio': date.year,
        'cveoportunidad': link.pgla,
        'fecha': date.strftime("%d-%m-%Y %H:%M:%S"),
        'observaciones': 'reactivated',
        'command': '1',
    }

    html = getHTMLContentFromPGLA(url, method='post', username='ldiaz', password='12345', data=data)

def generatePDF(link):
    pdfPath = os.path.join('CN', 'COMPLETION_NOTICE_' + link.client.name + '_PGLA_' + str(link.pgla) + '_NSR_' + link.nsr + '.pdf')
    pdf = FPDF()
    pdf.add_page()
    pdf.ln(0)
    pdf.set_font('Arial', '', 10)
    pdf.cell(40, 10, 'Hello World!')
    #pdf.output('tuto1.pdf', 'F')

    #pdf = FPDF()
    #pdf.add_page()
    #pdf.set_font('Arial', 'B', 16)
    #pdf.cell(40, 10, 'Hello World!')
    #pdf.cell(0, 10, 'NSR/CIRCUIT ID:' + project.nsr, 0, ln=1)
    #pdf.ln(0)
    #pdf.cell(0, 10, 'Billing Effective:' + project.billing_date, 0, 1)
    #pdf.cell(0, 10, 'Location:' + project.address, 0, ln=1)
    byte_string = pdf.output(dest="S")
    return BytesIO(byte_string)


def generateInventoryReportEXCEL(links, report):

    output = StringIO.StringIO()
    workbook = xlsxwriter.Workbook(output, {'strings_to_urls': False})
    format1 = workbook.add_format(
        {'text_wrap': True, 'bold': 1, 'border': 1, 'border_color': 'black', 'bg_color': 'blue', 'font_color': 'white',
         'align': 'center'})
    format2 = workbook.add_format({'text_wrap': True, 'border': 1, 'border_color': 'black'})

    worksheet = workbook.add_worksheet("Database")
    header = []
    for field in report:
        field = field.upper().replace('_', ' ')
        header.append(field)

    for column, value in enumerate(header):
        worksheet.write(0, column, value, format1)

    for row, link in enumerate(links):
        for i, field in enumerate(report):
            if field in ['client', 'config']:
                worksheet.write(row + 1, i, getattr(link, field).__repr__(), format2)
            else:
                worksheet.write(row + 1, i, getattr(link, field), format2)

    workbook.close()

    output.seek(0)

    return output



#downloadConfiguration()
#updateDatabase('MXXXM0094178-4')
#readEmails()

#createRFS("42152", "MXXXM60464-6", link_data)
#print(getSpeedInterfaceProfile("42152", "MXXXM60464-6", "7"))
#print(getDataFromPGLA("MXXXM0095051-9"))
#print(data_alta)
#print(get_config('vpn-coa-ayuntamiento-12', 'show running-config'))
#downloadBrasilConfigs()
#downloadMexicioConfigs()
#send_mail('guillermo.diaz@telmex.com', ['guillermo.diaz@telmex.com'], 'Prueba', 'Prueba')
