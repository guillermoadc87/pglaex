import xlsxwriter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from io import BytesIO
from django.http import StreamingHttpResponse
from wsgiref.util import FileWrapper

report_state = {
    'INSTALACION SUSPENDIDA': 'CUSTOMER HOLD',
    'ACCESO SOLICITADO (ACSO)': 'PROVISIONING W/FOC',
    'ACCESO LISTO (ACLI)': 'PROVISIONING W/FOC',
    'DESCONEXION SOLICITADA (DXSO)': 'PROVISIONING W/FOC',
    'DESCONEXION LISTA (DXLI)': 'PROVISIONING W/FOC',
    'PRUEBAS CON EL CLIENTE': 'COMPLETED',
    'ACTIVO SIN FACTURACION': 'COMPLETED',
}

report_types = {
    'RPV Multiservicios': 'MPLS',
    'INTERNET': 'DIA'
}

report_change_type = {
    'CAMBIO': 'L',
    'CAMBIO FISICO': 'F'
}

report_movement = {
    'ALTA': 'NEW',
    'CAMBIO': 'CHANGE',
    'CAMBIO FISICO': 'CHANGE'
}

def duplicate_service(modeladmin, request, queryset):
    for link in queryset:
        link.pk = None
        link.save()
duplicate_service.short_description = "Duplicate selected services"

def all_days_report(modeladmin, request, queryset):

    if queryset:
        output = BytesIO()

        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        bd = Side(style='thick', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        wb = xlsxwriter.Workbook(output)
        format1 = wb.add_format(
            {'text_wrap': True, 'bold': 1, 'border': 1, 'border_color': 'black', 'bg_color': 'blue',
             'font_color': 'white', 'align': 'center'})
        format2 = wb.add_format({'text_wrap': True, 'border': 1, 'border_color': 'black'})
        ws = wb.add_worksheet("Report")

        ws.set_column('A:B', 30)
        ws.set_column('D:D', 30)

        ws.write(0, 0, 'STATUS', format1)
        ws.write(0, 1, 'SITE NAME', format1)
        ws.write(0, 2, 'PGLA', format1)
        ws.write(0, 3, 'NSR', format1)
        ws.write(0, 4, 'TYPE', format1)
        ws.write(0, 5, 'EORDER DAYS', format1)
        ws.write(0, 6, 'LOCAL ORDER DAYS', format1)
        ws.write(0, 7, 'TOTAL', format1)
        ws.write(0, 8, 'CNR', format1)
        ws.write(0, 9, 'CYCLE TIME', format1)
        for row, link in enumerate(queryset):
            ws.write(row+1, 0, report_state.get(link.state, 'None'), format2)
            ws.write(row + 1, 1, link.site_name, format2)
            ws.write(row + 1, 2, link.pgla, format2)
            ws.write(row + 1, 3, link.nsr, format2)
            ws.write(row + 1, 4, link.movement.name, format2)
            ws.write(row + 1, 5, link.eorder_days, format2)
            ws.write(row + 1, 6, link.local_order_days, format2)
            ws.write(row + 1, 7, link.total, format2)
            ws.write(row + 1, 8, link.cnr, format2)
            ws.write(row + 1, 9, link.cycle_time, format2)

        wb.close()

        output.seek(0)

        response = StreamingHttpResponse(FileWrapper(output), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = "attachment; filename=report.xlsx"

        return response
all_days_report.short_description = "All Days Report"

def ct_report(modeladmin, request, queryset):

    if queryset:
        output = BytesIO()

        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        bd = Side(style='thick', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        wb = xlsxwriter.Workbook(output)
        format1 = wb.add_format(
            {'text_wrap': True, 'bold': 1, 'border': 1, 'border_color': 'black', 'bg_color': '#00B0F0',
             'align': 'center', 'font_size': 10, 'font_name': 'Calibri'})
        format2 = wb.add_format({'text_wrap': True, 'border': 1, 'border_color': 'black', 'font_size': 10, 'font_name': 'Calibri'})
        ws = wb.add_worksheet("Report")

        ws.set_column('A:P', 15)
        ws.set_column('D:D', 23)
        ws.set_column('H:H', 20)

        ws.write(0, 0, 'CUSTOMER', format1)
        ws.write(0, 1, 'PM', format1)
        ws.write(0, 2, 'PGLA', format1)
        ws.write(0, 3, 'NSR', format1)
        ws.write(0, 4, 'Local ID', format1)
        ws.write(0, 5, 'ORDER TYPE', format1)
        ws.write(0, 6, 'SERVICE', format1)
        ws.write(0, 7, 'COUNTRY', format1)
        ws.write(0, 8, 'PM ASSIGN', format1)
        ws.write(0, 9, 'CIAP DUE DATE', format1)
        ws.write(0, 10, 'BILLING DATE', format1)
        ws.write(0, 11, 'CNR Days', format1)
        ws.write(0, 12, 'CT', format1)
        ws.write(0, 13, 'OTP', format1)
        ws.write(0, 14, 'ADJPMA', format1)
        ws.write(0, 15, 'TYPE OF CHANGE', format1)
        for row, link in enumerate(queryset):
            ws.write(row + 1, 0, link.customer.name, format2)
            ws.write(row + 1, 1, link.pm, format2)
            ws.write(row + 1, 2, link.pgla, format2)
            ws.write(row + 1, 3, link.nsr, format2)
            ws.write(row + 1, 4, link.local_id, format2)
            ws.write(row + 1, 5, report_movement.get(link.movement.name, link.movement.name), format2)
            ws.write(row + 1, 6, report_types.get(link.service, link.service), format2)
            ws.write(row + 1, 7, link.country.name, format2)
            ws.write(row + 1, 8, link.reception_ciap.strftime("%m/%d/%Y") if link.reception_ciap else 'None', format2)
            ws.write(row + 1, 9, link.duedate_ciap.strftime("%m/%d/%Y") if link.duedate_ciap else 'None', format2)
            ws.write(row + 1, 10, link.billing_date.strftime("%m/%d/%Y") if link.billing_date else 'None', format2)
            ws.write(row + 1, 11, link.cnr, format2)
            ws.write(row + 1, 12, link.cycle_time, format2)
            ws.write(row + 1, 13, 1 if link.otp else 0, format2)
            ws.write(row + 1, 14, link.adjusted_due_date.strftime("%m/%d/%Y") if link.adjusted_due_date else 'None', format2)
            ws.write(row + 1, 15, report_change_type.get(link.movement.name, 'N/A'), format2)

        wb.close()

        output.seek(0)

        response = StreamingHttpResponse(FileWrapper(output), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = "attachment; filename=report.xlsx"

        return response
ct_report.short_description = "CT Report"